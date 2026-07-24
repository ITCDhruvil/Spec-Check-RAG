"""
Analytics export: preview data (JSON) + Excel workbook.

Sheets / tabs:
  1. Summary    — team (or user) totals for the selected period
  2. Per User   — one row per user with activity metrics
  3. Documents  — transposed matrix: one ROW per spec field, one COLUMN GROUP
                  per document with three sub-columns (Status / Correction /
                  Reason). Wrong fields carry the user's corrected value and
                  why it was wrong (issue type + comment).

Scoping: general users see only their own documents (and only their own row
in Per User); management may scope to one user or the whole team.
"""

from __future__ import annotations

from datetime import timedelta
from io import BytesIO
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.documents.models import Document
from apps.intelligence.models import FieldFeedback, GeneratedSummary
from apps.intelligence.services.spec_check_fields_registry import (
    DEADLINE_LABEL_DISPLAY,
    FIELD_DEFS,
)
from apps.intelligence.services.user_insights_service import build_user_insights

HEADER_FILL = PatternFill("solid", fgColor="2E5E8C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF1F7")
SUBHEADER_FONT = Font(bold=True, size=9, color="2E5E8C")
CORRECT_FILL = PatternFill("solid", fgColor="E7F3EC")
WRONG_FILL = PatternFill("solid", fgColor="FDE8E8")
NOT_FOUND_FONT = Font(color="9AA0A8", italic=True, size=10)
CELL_FONT = Font(size=10)

MAX_EXPORT_DOCS = 100


# Events (contextual notes) attached to date / acquisition rows — exported
# as their own rows when present.
EVENT_FIELDS: list[tuple[str, str]] = [
    ("project_document_acquisition_note", "Document acquisition — Events"),
    ("bid_deadline_date_time", "Bid deadline — Events"),
    ("bid_open_date_time", "Bid open — Events"),
    ("pre_bid_deadline_date_time", "Pre-bid — Events"),
    ("site_visit_date_time", "Site visit — Events"),
    ("question_deadline_date_time", "Question deadline — Events"),
    ("municipal_meeting_date_time", "Award — Events"),
]


def _canonical_fields() -> list[tuple[str, str]]:
    """(field_key, display label) for every spec-check field, stable order."""
    fields = [(name, d.display_label) for name, d in FIELD_DEFS.items()]
    fields += [(k, v) for k, v in DEADLINE_LABEL_DISPLAY.items()]
    fields.append(("set_aside", "Set-aside"))
    seen: set[str] = set()
    out: list[tuple[str, str]] = []
    for key, label in fields:
        if key not in seen:
            seen.add(key)
            out.append((key, label))
    return out


def build_export_data(
    days: int = 30,
    *,
    scope_user_id: str | None = None,
) -> dict[str, Any]:
    """All three tabs as JSON — used by the preview page and the xlsx writer."""
    since = timezone.now() - timedelta(days=days) if days else None

    insights = build_user_insights(days=days)
    users = insights["users"]
    if scope_user_id:
        users = [u for u in users if u["user_id"] == str(scope_user_id)]

    totals_docs = sum(u["docs_total"] for u in users)
    totals_completed = sum(u["docs_completed"] for u in users)
    totals_failed = sum(u["docs_failed"] for u in users)
    totals_fields = sum(u["fields_extracted"] for u in users)
    totals_corrected = sum(u["fields_corrected"] for u in users)

    summary_rows = [
        ("Period", f"Last {days} days" if days else "All time"),
        ("Generated", timezone.now().strftime("%B %d, %Y %I:%M %p")),
        ("Scope", users[0]["username"] if scope_user_id and users else "All users"),
        ("Active users", len(users)),
        ("Documents uploaded", totals_docs),
        ("Documents completed", totals_completed),
        ("Documents failed", totals_failed),
        ("Fields extracted", totals_fields),
        ("Fields corrected (thumbs down)", totals_corrected),
        (
            "Overall correction rate",
            f"{round(totals_corrected / totals_fields * 100, 1)}%" if totals_fields else "0%",
        ),
    ]

    # ── Documents (transposed): fields as rows, documents as column groups ──
    docs_qs = Document.objects.select_related("uploaded_by").order_by("-created_at")
    if since:
        docs_qs = docs_qs.filter(created_at__gte=since)
    if scope_user_id:
        docs_qs = docs_qs.filter(uploaded_by_id=scope_user_id)
    docs = list(docs_qs[:MAX_EXPORT_DOCS])
    doc_ids = [d.id for d in docs]

    # Final user-facing fields per document (current completed summary),
    # plus each date row's Events note text.
    extracted: dict[str, set[str]] = {}
    event_notes: dict[str, dict[str, str]] = {}
    for doc_id, sjson in GeneratedSummary.objects.filter(
        document_id__in=doc_ids, is_current=True, status="completed"
    ).values_list("document_id", "summary_json"):
        labels = extracted.setdefault(str(doc_id), set())
        notes = event_notes.setdefault(str(doc_id), {})
        spec = (sjson or {}).get("spec_check_fields") or {}
        for bucket in spec.values():
            if not isinstance(bucket, list):
                continue
            for row in bucket:
                if isinstance(row, dict):
                    fk = str(row.get("field_key") or "").strip()
                    if fk:
                        labels.add(fk)
                        note = str(row.get("_note") or "").strip()
                        if note:
                            notes[fk] = note

    # Wrong fields with correction + reason per (document, field).
    wrong: dict[tuple[str, str], dict[str, str]] = {}
    for fb in FieldFeedback.objects.filter(
        document_id__in=doc_ids, rating="down"
    ).values("document_id", "field_key", "correct_value", "issue_type", "comment"):
        key = (str(fb["document_id"]), fb["field_key"])
        reason_parts = [p for p in (fb["issue_type"], fb["comment"]) if p]
        # Keep the richest feedback per field (prefer entries with a correction).
        prior = wrong.get(key)
        if prior and prior["correction"] and not fb["correct_value"]:
            continue
        wrong[key] = {
            "correction": fb["correct_value"] or "",
            "reason": " — ".join(reason_parts) if reason_parts else "Marked wrong",
        }

    fields = _canonical_fields()
    doc_columns = [
        {
            "id": str(d.id),
            "filename": d.original_filename,
            "uploaded_by": d.uploaded_by.username if d.uploaded_by else "",
            "status": d.status,
            "uploaded": d.created_at.strftime("%Y-%m-%d"),
        }
        for d in docs
    ]

    field_rows: list[dict[str, Any]] = []
    for key, label in fields:
        cells = []
        for d in docs:
            did = str(d.id)
            w = wrong.get((did, key))
            if w is not None:
                cells.append({"status": "Wrong", **w})
            elif key in extracted.get(did, set()):
                cells.append({"status": "Correct", "correction": "", "reason": ""})
            else:
                cells.append({"status": "Not found", "correction": "", "reason": ""})
        field_rows.append({"field_key": key, "label": label, "cells": cells})

    # Events rows: the note text itself (in the Correction column so the full
    # text is visible; Status marks presence).
    for key, label in EVENT_FIELDS:
        cells = []
        any_note = False
        for d in docs:
            note = event_notes.get(str(d.id), {}).get(key, "")
            if note:
                any_note = True
                cells.append({"status": "Correct", "correction": note, "reason": ""})
            else:
                cells.append({"status": "Not found", "correction": "", "reason": ""})
        if any_note:
            field_rows.append(
                {"field_key": f"{key}_events", "label": label, "cells": cells}
            )

    return {
        "period_days": days,
        "scope_user_id": str(scope_user_id) if scope_user_id else None,
        "summary": [{"metric": m, "value": str(v)} for m, v in summary_rows],
        "per_user": users,
        "documents": {"columns": doc_columns, "rows": field_rows},
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_analytics_workbook(
    days: int = 30, *, scope_user_id: str | None = None
) -> bytes:
    data = build_export_data(days, scope_user_id=scope_user_id)
    wb = Workbook()

    # 1 ── Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(("Metric", "Value"))
    for row in data["summary"]:
        ws.append((row["metric"], row["value"]))
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = CELL_FONT
    _style_header_row(ws, 2)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30

    # 2 ── Per User
    ws = wb.create_sheet("Per User")
    headers = (
        "Username", "Email", "Documents", "Completed", "Failed",
        "Fields extracted", "Fields corrected", "Correction rate %",
        "Avg processing (s)", "Last activity",
    )
    ws.append(headers)
    for u in data["per_user"]:
        ws.append((
            u["username"], u["email"], u["docs_total"], u["docs_completed"],
            u["docs_failed"], u["fields_extracted"], u["fields_corrected"],
            u["correction_rate"], u["avg_processing_seconds"],
            (u["last_activity"] or "")[:10],
        ))
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = CELL_FONT
    _style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # 3 ── Documents (transposed: field rows × document column groups)
    ws = wb.create_sheet("Documents")
    doc_cols = data["documents"]["columns"]
    rows = data["documents"]["rows"]

    # Row 1: document names (merged over each 3-column group)
    ws.cell(row=1, column=1, value="Field")
    for i, doc in enumerate(doc_cols):
        start = 2 + i * 3
        ws.cell(row=1, column=start, value=doc["filename"])
        ws.merge_cells(
            start_row=1, start_column=start, end_row=1, end_column=start + 2
        )
    _style_header_row(ws, 1 + len(doc_cols) * 3)

    # Row 2: sub-columns
    ws.cell(row=2, column=1, value="")
    for i in range(len(doc_cols)):
        start = 2 + i * 3
        for offset, sub in enumerate(("Status", "Correction", "Reason")):
            cell = ws.cell(row=2, column=start + offset, value=sub)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
    ws.freeze_panes = "B3"

    # Field rows
    for r_idx, frow in enumerate(rows, start=3):
        name_cell = ws.cell(row=r_idx, column=1, value=frow["label"])
        name_cell.font = Font(bold=True, size=10)
        for i, cell_data in enumerate(frow["cells"]):
            start = 2 + i * 3
            status_cell = ws.cell(row=r_idx, column=start, value=cell_data["status"])
            corr_cell = ws.cell(row=r_idx, column=start + 1, value=cell_data["correction"])
            reason_cell = ws.cell(row=r_idx, column=start + 2, value=cell_data["reason"])
            corr_cell.font = CELL_FONT
            reason_cell.font = CELL_FONT
            if cell_data["status"] == "Correct":
                status_cell.fill = CORRECT_FILL
                status_cell.font = CELL_FONT
            elif cell_data["status"] == "Wrong":
                status_cell.fill = WRONG_FILL
                status_cell.font = CELL_FONT
                corr_cell.fill = WRONG_FILL
                reason_cell.fill = WRONG_FILL
            else:
                status_cell.font = NOT_FOUND_FONT

    ws.column_dimensions["A"].width = 30
    for i in range(len(doc_cols)):
        start = 2 + i * 3
        ws.column_dimensions[get_column_letter(start)].width = 11
        ws.column_dimensions[get_column_letter(start + 1)].width = 22
        ws.column_dimensions[get_column_letter(start + 2)].width = 26

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
